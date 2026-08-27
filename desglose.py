"""
Despiece (lista de materiales) multinivel de un articulo del ERP GOMEZYCRESPO.

Explota TODOS los niveles con un CTE recursivo, multiplicando las cantidades
en cascada, y marca que lineas son materia prima (hoja del arbol).

Uso:
    py desglose.py 12101021
    py desglose.py 12101021 -o salidas/mi_desglose.xlsx
"""
import argparse
import os
import re

import pandas as pd
from openpyxl.styles import PatternFill, Font
from sqlalchemy import text

from db import get_engine

# CTE recursivo. Costeo de compras (criterio: TIPO DE APROVISIONAMIENTO):
#   - se despieza TODO lo que tiene escandallo, a todos los niveles, por DOS vias:
#       (1) FASE activa (fabricacion) -> Fases_Entradas.
#       (2) CONJUNTO (Articulos_Conjuntos) -> SOLO si el articulo no tiene fase activa.
#     Un conjunto sin fase se explota por sus componentes x Unidades (misma logica
#     recursiva que las fases). Prioridad: si hay fase activa manda la fase.
#   - se valora a ultimo precio de compra cada articulo HOJA del arbol, y se
#     clasifica por su tipo (MATERIA PRIMA / MERCADERIA / ...).
#   - un articulo que TIENE escandallo Y ADEMAS se compra tambien se valora, y su
#     precio se SUMA al de sus componentes: esa compra es el trabajo que se le
#     hace encima (lacado, zincado, remachado). El CUENCO LACADO 18104023 cuesta
#     6,50 de cuenco en bruto + 1,05 de lacado = 7,55.
#     SALVAVIDAS: si ese precio ya esta dentro del escandallo (un componente que
#     cuesta lo mismo porque es el articulo comprado hecho) NO se suma, que si no
#     se cuenta dos veces. Ver el CTE 'duplicado'.
# Articulos que NO deben sumar su precio propio y que el freno automatico (2%)
# NO detecta, porque su version equivalente tiene un precio bastante distinto.
# Van a mano porque no hay regla de datos que los separe de un lacado legitimo:
# 18104044 (CUENCO V,2 lacado) cuesta 2,50 sobre una pieza de 2,62 -> 4,8% de
# desvio y es correcto sumarlo, mientras que 18201029 desvia un 6,3% y no lo es.
# Cualquier banda que atrape a uno se lleva al otro por delante.
NO_SUMA_PROPIA = {
    "10701016",  # MANGUERA TROZO DE 15CM (0,0598): comprar el trozo cortado o
                 # cortarlo de 10701003 (manguera por metros) son ALTERNATIVAS
    "18201028",  # MASTIL HAMACA GATO LACADO 1,50 vs MASTIL sin lacar 1,30
    "18201029",  # BASE HAMACA GATO LACADO  1,60 vs BASE   sin lacar 1,50
    "22306070",  # CUADRO SINFIN Monofasico (600) lleva dentro el TRIFASICO (578):
                 # son productos distintos, el escandallo esta mal montado
    "11601001",  # BEBEDERO AUTOMATICO COLGANTE (8,30) lleva dentro 11601067
                 # BEBEDERO AUTOMATICO Arion (8,90): el mismo bebedero comprado
                 # en fechas distintas -> 6,7% de desvio, el freno no lo pilla
}
_NO_SUMA_LISTA = ", ".join("'%s'" % c for c in sorted(NO_SUMA_PROPIA))
_NO_SUMA_VALUES = ", ".join("('%s')" % c for c in sorted(NO_SUMA_PROPIA))

_SQL = """
WITH
-- Precio de compra por articulo, con prioridad:
--   1) ULTIMO ALBARAN con precio real (Pedidos_Prov_Lineas, la compra mas reciente).
--   2) FALLBACK: TARIFA del proveedor (Listas_Precios_Prov_Art), solo si no hay albaran.
--
-- La tarifa es lo que el proveedor PIDE, no lo que se ha pagado, asi que manda
-- siempre el albaran. Pero la lista de precios es un canal vivo: el equipo la
-- mantiene a diario (429 cambios en 2026, 5 usuarios), y hay articulos que solo
-- tienen precio ahi. Ignorarla dejaba a cero cosas con precio recien puesto.
-- Toda linea valorada por tarifa se marca con Fuente='lista' y sale señalada en
-- pantalla, para no confundir un precio de catalogo con una compra real.
--
-- OJO: tarifa y albaran pueden venir en UNIDADES distintas (p.ej. cordon a
-- 0,026 EUR/m en tarifa frente a 24,79 EUR/rollo en albaran). Al revisar un
-- precio que venga de lista, comprobar la unidad.
--
-- SE EXCLUYE LA LISTA DEL "PROVEEDOR 0" (IdLista 0). No es un proveedor: es el
-- hueco generico que creo la instalacion del ERP en 2001, con vigencia caducada
-- (HastaFecha 2021-01-01). Acumula 6.071 lineas, y buena parte son articulos que
-- GYC FABRICA (PISO VARILLA, LATERAL NIDO SPRINT, MODULO RODEIRO...), donde el
-- importe es una valoracion interna, no lo que pide un proveedor. Costear con eso
-- es inventarse un precio de compra que nadie ha ofertado.
-- Lineas con Descuento=100 se excluyen: son regalos/unidades gratis dentro de
-- un pedido (misma FechaAlbaran/IdPedido que la linea real, p.ej. GRAPADORA
-- MANUAL 25701001: 2 lineas del mismo pedido a 6,95, una al 0% y otra al 100%).
-- Cuando es la mas reciente, el ROW_NUMBER la elegia y el descuento la dejaba
-- en 0 EUR de coste sin avisar (Precio no es NULL -> no salta "Sin precio").
-- Hay 4 casos en toda la base; en 3 de ellos es la UNICA compra que existe, asi
-- que ahi no hay precio real que rescatar (cae a tarifa o a "sin precio", que
-- es lo correcto: mejor avisar que enseñar un precio que nunca se pago).
precio AS (
    SELECT IdArticulo, Precio, Descuento, 'albaran' AS Fuente FROM (
        SELECT IdArticulo, Precio_EURO AS Precio, Descuento,
               ROW_NUMBER() OVER (PARTITION BY IdArticulo
                                  ORDER BY FechaAlbaran DESC, IdPedido DESC) AS rn
        FROM dbo.Pedidos_Prov_Lineas
        WHERE FechaAlbaran IS NOT NULL AND Precio_EURO > 0 AND Descuento < 100
    ) t
    WHERE rn = 1

    UNION ALL

    -- de la tarifa se coge el precio MAS RECIENTE (por FechaInsertUpdate) de
    -- entre las listas de PROVEEDORES DE VERDAD (ver nota del "Proveedor 0")
    SELECT IdArticulo, Precio, 0 AS Descuento, 'lista' AS Fuente FROM (
        SELECT lp.IdArticulo, lp.Precio,
               ROW_NUMBER() OVER (PARTITION BY lp.IdArticulo
                                  ORDER BY lp.FechaInsertUpdate DESC, lp.IdLista DESC) AS rn
        FROM dbo.Listas_Precios_Prov_Art lp
        INNER JOIN dbo.Listas_Precios_Prov lc ON lc.IdLista = lp.IdLista
                                             AND lc.IdProveedor <> '0'
        WHERE lp.Precio > 0
          AND NOT EXISTS (SELECT 1 FROM dbo.Pedidos_Prov_Lineas pl
                          WHERE pl.IdArticulo = lp.IdArticulo
                            AND pl.FechaAlbaran IS NOT NULL AND pl.Precio_EURO > 0
                            AND pl.Descuento < 100)
    ) t
    WHERE t.rn = 1
),
-- MANO DE OBRA IMPUTADA A MANO (Trabajos_ManoObra). Es la fuente PREFERENTE de
-- tiempo: produccion la rellena operacion a operacion en el ERP, asi que cuando
-- existe manda sobre la media de bonos.
--   - Duracion viene en DIAS (IdUnidadDuracion = 'D'), de ahi el x1440 para
--     pasarla a minutos. Comprobado contra CosteTotal: 0,001389 D x 24 h x 17
--     EUR/h = 0,5667 EUR, que es exactamente lo que guarda la fila.
--   - se enlaza por Trabajos_Fases: Trabajos_ManoObra.IdTrabajo -> IdFase -> el
--     articulo que sale de esa fase.
--   - se SUMAN las lineas de una misma fase (una fase puede llevar varios
--     trabajos, cada uno con su mano de obra) y luego se agrupa por articulo.
--     El MAX final es defensivo: hoy ningun articulo sale de dos fases activas,
--     asi que con una sola fase es indiferente.
--   - Duracion = 0 NO es una medicion, es la casilla sin rellenar (mismo
--     criterio que con los precios a cero), asi que no cuenta y se cae al bono.
tiempo_mano_obra AS (
    SELECT fs.IdArticulo, MAX(m.Minutos) AS TiempoMin
    FROM (
        SELECT tf.IdFase, SUM(tmo.Duracion) * 1440.0 AS Minutos
        FROM dbo.Trabajos_ManoObra tmo
        INNER JOIN dbo.Trabajos_Fases tf ON tf.IdTrabajo = tmo.IdTrabajo
        WHERE tmo.Duracion > 0
        GROUP BY tf.IdFase
    ) m
    INNER JOIN dbo.Fases f5 ON f5.IdFase = m.IdFase AND f5.Activa <> 0
    INNER JOIN dbo.Fases_Salidas fs ON fs.IdFase = m.IdFase
    GROUP BY fs.IdArticulo
),
-- RESPALDO: media de los bonos de produccion (TotalMinutos/TotalPiezas) con fase
-- activa. Se usa solo si el articulo no tiene mano de obra imputada, y entonces
-- la linea sale marcada como "medio" en pantalla: es la media REAL de lo
-- que se tardo, no un tiempo teorico. El teorico seria el estandar de
-- produccion, que es justo la otra via (mano de obra imputada).
tiempo_op AS (
    SELECT obs.IdArticulo,
           AVG(CAST(ob.TotalMinutos AS float) / NULLIF(ob.TotalPiezas, 0)) AS TiempoMin
    FROM dbo.Ordenes_Bonos_Salidas obs
    INNER JOIN dbo.Ordenes_Articulos oa ON oa.IdArticulo = obs.IdArticulo AND oa.IdOrden = obs.IdOrden
    INNER JOIN dbo.Fases f ON f.IdFase = oa.IdFase AND f.Activa <> 0
    INNER JOIN dbo.Ordenes_Bonos ob ON ob.IdBono = obs.IdBono AND ob.IdOrden = obs.IdOrden
    WHERE ob.TotalPiezas > 0
    GROUP BY obs.IdArticulo
),
-- Articulos cuya fase activa NO declara ninguna operacion de fabricacion:
-- todos sus trabajos tienen OrdenTrabajo = 0 (los "Sin operacion" del ERP).
-- La fase existe solo para llevar los materiales. Que no tengan tiempo NO es
-- un hueco de datos, es lo correcto: no hay mano de obra que imputar.
-- Se exige que NO haya ningun trabajo con OrdenTrabajo = 1 (seguro por si
-- aparecen fases mixtas; hoy no hay ninguna en la base).
--
-- LA CASILLA OrdenTrabajo ES EL UNICO CRITERIO. Es el campo que produccion
-- repasa en el ERP, asi que manda lo que diga: para quitar el aviso de una
-- pieza que no lleva mano de obra, basta con desmarcarla ahi.
--
-- Antes se exigia ademas que el trabajo se LLAMASE "sin operacion". Eso dejaba
-- fuera 31 articulos con la casilla desmarcada pero nombre de operacion real
-- (Lacar x16, Cortar, Soldar, Punzonar, Inyectar, Estampar...), que salian
-- avisados como "sin tiempo". Ahora entran aqui y pasan a 0 EUR de mano de
-- obra EN SILENCIO. Es intencionado: el aviso se sustituye por el repaso
-- manual del ERP. Si aparece un 0 EUR de operacion sin explicacion, mirar
-- aqui primero.
sin_operacion AS (
    SELECT DISTINCT fs.IdArticulo
    FROM dbo.Fases_Salidas fs
    INNER JOIN dbo.Fases f ON f.IdFase = fs.IdFase AND f.Activa <> 0
    WHERE EXISTS (SELECT 1 FROM dbo.Trabajos_Fases tf
                  WHERE tf.IdFase = fs.IdFase AND tf.OrdenTrabajo = 0)
      AND NOT EXISTS (SELECT 1 FROM dbo.Trabajos_Fases tf
                      WHERE tf.IdFase = fs.IdFase AND tf.OrdenTrabajo = 1)
),
-- Componentes de cada articulo, unificando las dos vias de escandallo.
componentes AS (
    -- (1) via FASE ACTIVA (fabricacion)
    SELECT
        fs.IdArticulo AS Padre,
        fe.IdArticulo AS Hijo,
        fe.Descrip    AS Descripcion,
        fe.IdTipoUnidad AS Unidad,
        TRY_CONVERT(FLOAT, fe.Cantidad) AS Cantidad
    FROM dbo.Fases f
    INNER JOIN dbo.Fases_Salidas  fs ON fs.IdFase = f.IdFase AND f.Activa <> 0
    INNER JOIN dbo.Fases_Entradas fe ON fe.IdFase = f.IdFase

    UNION ALL

    -- (2) via CONJUNTO, SOLO si el articulo padre NO tiene fase activa
    SELECT
        ac.IdArticuloPadre AS Padre,
        ac.IdArticulo      AS Hijo,
        a.Descrip          AS Descripcion,
        CAST(NULL AS VARCHAR(20)) AS Unidad,
        TRY_CONVERT(FLOAT, ac.Unidades) AS Cantidad
    FROM dbo.Articulos_Conjuntos ac
    INNER JOIN dbo.Articulos a ON a.IdArticulo = ac.IdArticulo
    WHERE NOT EXISTS (SELECT 1 FROM dbo.Fases_Salidas fs2
                      INNER JOIN dbo.Fases f2 ON f2.IdFase = fs2.IdFase AND f2.Activa <> 0
                      WHERE fs2.IdArticulo = ac.IdArticuloPadre)
      -- corta ciclos DEGENERADOS de 2 nodos: el conjunto tiene un UNICO
      -- componente y es precisamente el mismo articulo que ya lo lleva a EL
      -- dentro por su fase activa (18608010 lleva dentro 60009006 por su fase,
      -- y ademas 60009006 esta metido como "conjunto" formado por 1 unidad de
      -- 18608010 -> bucle que hacia desaparecer el precio del NIDO...(comprado)
      -- sin avisar). Hay 152 conjuntos con este mismo bucle fase<->conjunto en
      -- toda la base, pero en 148 el conjunto agrupa 2 o mas piezas REALES
      -- (p.ej. CONJUNTO GANCHOS = gancho izq. + gancho der.): ahi la relacion
      -- de conjunto es legitima y hay que dejarla, el bucle viene de que ADEMAS
      -- cada pieza referencia al conjunto en su propia fase (dato raro, pero no
      -- rompe el coste porque el conjunto no tiene precio propio que duplicar).
      -- Solo los 4 conjuntos de 1 solo componente son el caso degenerado real:
      -- 60008007, 60009006, 60009007, 60009009.
      AND (
            NOT EXISTS (SELECT 1 FROM dbo.Fases f3
                        INNER JOIN dbo.Fases_Salidas  fs3 ON fs3.IdFase = f3.IdFase AND f3.Activa <> 0
                        INNER JOIN dbo.Fases_Entradas fe3 ON fe3.IdFase = f3.IdFase
                        WHERE fs3.IdArticulo = ac.IdArticulo
                          AND fe3.IdArticulo = ac.IdArticuloPadre)
            OR (SELECT COUNT(*) FROM dbo.Articulos_Conjuntos ac9
                WHERE ac9.IdArticuloPadre = ac.IdArticuloPadre) <> 1
          )
),
-- Articulos cuyo PRECIO PROPIO YA ESTA DENTRO de su escandallo: alguno de sus
-- componentes directos cuesta lo mismo que ellos, porque es el mismo articulo
-- comprado hecho (BEBEDERO DT20 91,00 lleva dentro "BEBEDERO DT20 ... comprado"
-- a 91,00; ABREVADERO 78,00 lleva "ABREVADERO ... (comprado)" a 78,00).
-- Ahi el precio propio y el escandallo describen LO MISMO por dos caminos, y
-- sumarlos duplica: el bebedero pasaria de 91 a 182 EUR.
--
-- Son 6 en toda la base (CUADRO ROBOT, BEBEDERO DT20, ABREVADERO, COMEDERO
-- ESQUINA, BOTELLA GYC, TOLVA DE PLASTICO). Los otros 58 llevan dentro la pieza
-- EN BRUTO, no el articulo terminado, y su precio propio es trabajo anadido
-- (lacado, zincado, remachado...) que SI hay que sumar.
duplicado AS (
    SELECT DISTINCT c.Padre AS IdArticulo
    FROM componentes c
    INNER JOIN precio ph ON ph.IdArticulo = c.Hijo
    INNER JOIN precio pp ON pp.IdArticulo = c.Padre
    WHERE ABS(ph.Precio - pp.Precio) <= 0.02 * ABS(pp.Precio)

    UNION

    -- los revisados a mano (ver NO_SUMA_PROPIA)
    SELECT v.IdArticulo FROM (VALUES /*NO_SUMA_VALUES*/) v(IdArticulo)
),
desglose AS (
    -- NIVEL 0: componentes directos del articulo (por fase o por conjunto)
    SELECT
        CAST(:codigo AS VARCHAR(50)) AS Articulo,
        c.Hijo        AS IdArticulo,
        c.Cantidad    AS Cant,
        c.Descripcion AS Componente,
        0 AS Nivel,
        c.Unidad AS Unidad,
        CAST('|' + CAST(:codigo AS VARCHAR(50)) + '|' + c.Hijo + '|' AS VARCHAR(4000)) AS Ruta
    FROM componentes c
    WHERE c.Padre = :codigo

    UNION ALL

    -- NIVEL 0 alternativo: el PROPIO articulo si NO tiene escandallo (compra directa).
    -- Se muestra a si mismo (cantidad 1) y se valora a su ultimo precio de compra.
    SELECT
        CAST(:codigo AS VARCHAR(50)) AS Articulo,
        a.IdArticulo AS IdArticulo,
        CAST(1 AS FLOAT) AS Cant,
        a.Descrip AS Componente,
        0 AS Nivel,
        CAST(NULL AS VARCHAR(20)) AS Unidad,
        CAST('|' + a.IdArticulo + '|' AS VARCHAR(4000)) AS Ruta
    FROM dbo.Articulos a
    WHERE a.IdArticulo = :codigo
      AND NOT EXISTS (SELECT 1 FROM componentes c WHERE c.Padre = :codigo)

    UNION ALL

    -- NIVELES SIGUIENTES: explota cada componente por su propio escandallo
    SELECT
        d.IdArticulo,
        c.Hijo,
        d.Cant * c.Cantidad,           -- cantidad acumulada (x Unidades en conjuntos)
        c.Descripcion,
        d.Nivel + 1,
        c.Unidad,
        CAST(d.Ruta + c.Hijo + '|' AS VARCHAR(4000))
    FROM desglose d
    INNER JOIN componentes c ON c.Padre = d.IdArticulo
    WHERE d.Ruta NOT LIKE '%|' + c.Hijo + '|%'          -- corta ciclos A->B->A
),
-- Marca cada fila: cantidad convertida (gr->kg), si es hoja (sin escandallo) y si
-- se puede valorar a precio de compra.
marcado AS (
    SELECT d.*,
        CASE WHEN d.Unidad = 'gr' THEN d.Cant / 1000.0 ELSE d.Cant END AS CantConv,
        CASE WHEN EXISTS (SELECT 1 FROM componentes c2 WHERE c2.Padre = d.IdArticulo)
             THEN 0 ELSE 1 END AS EsHoja,
        -- Valorable a precio de compra. Tres vias:
        --  (a) tiene TIPO DE APROVISIONAMIENTO (=alguien confirmo que se compra)
        --      y es HOJA del arbol (no tiene escandallo que explotar).
        --  (b) NO tiene tipo pero tampoco ha tenido NUNCA fase de fabricacion:
        --      solo puede ser una compra a la que le falta rellenar el tipo en el
        --      ERP. Se valora igual (si hay albaran) en vez de contarla como 0 EUR
        --      en silencio. Las piezas con fases desactivadas quedan fuera a
        --      proposito: esas siguen avisando como "sin escandallo".
        CASE WHEN a2.IdTipoAprovisionamiento IS NOT NULL
                  AND NOT EXISTS (SELECT 1 FROM componentes c3 WHERE c3.Padre = d.IdArticulo)
             THEN 1
             WHEN a2.IdTipoAprovisionamiento IS NULL
                  AND NOT EXISTS (SELECT 1 FROM componentes c4 WHERE c4.Padre = d.IdArticulo)
                  AND NOT EXISTS (SELECT 1 FROM dbo.Fases_Salidas fs3
                                  WHERE fs3.IdArticulo = d.IdArticulo)
             THEN 1
             -- (c) TIENE escandallo pero ADEMAS se compra: esa compra es el
             --     trabajo que se le hace encima (lacado, zincado, remachado),
             --     asi que se suma a sus componentes. Salvo que el precio ya
             --     este dentro del escandallo (ver CTE 'duplicado').
             WHEN EXISTS (SELECT 1 FROM componentes c5 WHERE c5.Padre = d.IdArticulo)
                  AND NOT EXISTS (SELECT 1 FROM duplicado dp WHERE dp.IdArticulo = d.IdArticulo)
             THEN 1
             ELSE 0 END AS EsValorable
    FROM desglose d
    LEFT JOIN dbo.Articulos a2 ON a2.IdArticulo = d.IdArticulo
)
SELECT
    m.Nivel,
    m.Articulo,
    m.IdArticulo,
    m.Componente,
    m.CantConv AS Cant,
    m.Unidad,
    t.Descrip AS TipoCompra,                       -- MATERIA PRIMA / MERCADERIA / ...
    -- Precio: ultimo ALBARAN, o TARIFA si no hay ninguna compra. Solo en las
    -- filas valorables (ver EsValorable en 'marcado').
    CASE WHEN m.EsValorable = 1 THEN p.Precio END    AS PrecioCompra,
    CASE WHEN m.EsValorable = 1 THEN p.Descuento END AS Descuento,
    CASE WHEN m.EsValorable = 1 AND p.Precio IS NOT NULL
         THEN p.Fuente END AS PrecioFuente,       -- 'albaran' o 'lista' (tarifa)
    CASE WHEN m.EsValorable = 1 AND p.Precio IS NOT NULL
         THEN m.CantConv * (p.Precio - p.Precio * COALESCE(p.Descuento, 0) / 100.0)
    END AS Coste,
    CASE WHEN m.EsValorable = 1 AND p.Precio IS NOT NULL THEN 1 ELSE 0 END AS Valorado,
    -- valorable pero sin precio en ninguna fuente: hueco de coste real
    CASE WHEN m.EsValorable = 1 AND p.Precio IS NULL
         THEN 1 ELSE 0 END AS SinPrecio,
    -- 1 si es una PIEZA FABRICADA (tiene fases) sin escandallo ACTIVO para
    -- costearla y sin valorar: su fase esta desactivada/vacia -> no se puede
    -- costear aqui. (NO marca tornilleria/compras: esas son sin-precio/sin-tipo.)
    CASE WHEN m.EsHoja = 1
              AND NOT (m.EsValorable = 1 AND p.Precio IS NOT NULL)
              AND EXISTS (SELECT 1 FROM dbo.Fases_Salidas fss WHERE fss.IdArticulo = m.IdArticulo)
         THEN 1 ELSE 0 END AS SinEscandallo,
    -- subconjunto de SinEscandallo que ademas es componente de un conjunto
    -- (su coste viene por el conjunto).
    CASE WHEN m.EsHoja = 1
              AND NOT (m.EsValorable = 1 AND p.Precio IS NOT NULL)
              AND EXISTS (SELECT 1 FROM dbo.Fases_Salidas fss WHERE fss.IdArticulo = m.IdArticulo)
              AND EXISTS (SELECT 1 FROM dbo.Articulos_Conjuntos ac WHERE ac.IdArticulo = m.IdArticulo)
         THEN 1 ELSE 0 END AS DeConjunto,
    -- min/pieza de operacion. Manda la mano de obra imputada; si no la hay, la
    -- media de bonos, y entonces TiempoMedio avisa de que es una estimacion.
    COALESCE(tmo.TiempoMin, tp.TiempoMin) AS TiempoOp,
    CASE WHEN tmo.TiempoMin IS NULL AND tp.TiempoMin IS NOT NULL
         THEN 1 ELSE 0 END AS TiempoMedio,
    -- 1 = su fase no declara operacion ("Sin operacion"): no lleva mano de obra
    -- y por tanto no debe avisarse como "sin tiempo".
    CASE WHEN so.IdArticulo IS NOT NULL THEN 1 ELSE 0 END AS SinOperacion,
    m.Ruta AS Ruta                                 -- ruta jerarquica para el arbol
FROM marcado m
LEFT JOIN dbo.Articulos  art ON art.IdArticulo = m.IdArticulo
LEFT JOIN precio         p   ON p.IdArticulo   = m.IdArticulo
LEFT JOIN tiempo_mano_obra tmo ON tmo.IdArticulo = m.IdArticulo
LEFT JOIN tiempo_op      tp  ON tp.IdArticulo  = m.IdArticulo
LEFT JOIN sin_operacion  so  ON so.IdArticulo  = m.IdArticulo
LEFT JOIN dbo.Articulos_Tipos_Aprovisionamiento t
       ON t.IdTipoAprovisionamiento = art.IdTipoAprovisionamiento
ORDER BY m.Ruta
OPTION (MAXRECURSION 0);   -- sin limite de niveles
"""

SQL = text(_SQL.replace("/*NO_SUMA_VALUES*/", _NO_SUMA_VALUES))


# Articulos.Estado: 0 = activo, 1 = dado de baja / descatalogado.
# (FechaBajaArt existe pero esta practicamente sin usar: 1 registro en toda la tabla,
#  asi que Estado es el unico indicador fiable de baja.)
SQL_BUSCAR = text("""
SELECT TOP (:limite)
    a.IdArticulo,
    a.Descrip,
    CASE WHEN EXISTS (
        SELECT 1 FROM dbo.Fases_Salidas fs
        INNER JOIN dbo.Fases f ON f.IdFase = fs.IdFase AND f.Activa <> 0
        WHERE fs.IdArticulo = a.IdArticulo
    ) OR EXISTS (
        SELECT 1 FROM dbo.Articulos_Conjuntos ac WHERE ac.IdArticuloPadre = a.IdArticulo
    ) THEN 1 ELSE 0 END AS Fabricable
FROM dbo.Articulos a
WHERE (a.IdArticulo LIKE :q OR a.Descrip LIKE :q)
  AND a.Estado = 0                             -- solo activos: los de baja no se buscan
ORDER BY Fabricable DESC, a.IdArticulo
""")


def buscar_articulos(q: str, limite: int = 50) -> pd.DataFrame:
    """Busca articulos ACTIVOS por codigo o descripcion. Los fabricables van
    primero. Los descatalogados (Estado = 1) no se devuelven nunca."""
    q = (q or "").strip()
    limite = max(1, min(int(limite), 200))
    with get_engine().connect() as cn:
        return pd.read_sql(SQL_BUSCAR, cn, params={"q": f"%{q}%", "limite": limite})


SQL_NOMBRE = text("SELECT Descrip FROM dbo.Articulos WHERE IdArticulo = :codigo")


def nombre_articulo(codigo: str) -> str:
    """Devuelve la descripcion del articulo (cadena vacia si no existe)."""
    codigo = re.sub(r"[^A-Za-z0-9]", "", str(codigo))
    with get_engine().connect() as cn:
        r = cn.execute(SQL_NOMBRE, {"codigo": codigo}).fetchone()
    return r[0] if r and r[0] else ""


# Tiempo del articulo RAIZ, con la misma prioridad que los CTE de arriba:
# manda la mano de obra imputada (Trabajos_ManoObra) y la media de bonos es el
# respaldo. Ver el comentario de tiempo_mano_obra para el porque del x1440.
SQL_TIEMPO = text("""
WITH mano_obra AS (
    SELECT SUM(tmo.Duracion) * 1440.0 AS TiempoMin
    FROM dbo.Trabajos_ManoObra tmo
    INNER JOIN dbo.Trabajos_Fases tf ON tf.IdTrabajo = tmo.IdTrabajo
    INNER JOIN dbo.Fases f2 ON f2.IdFase = tf.IdFase AND f2.Activa <> 0
    INNER JOIN dbo.Fases_Salidas fs ON fs.IdFase = tf.IdFase
    WHERE fs.IdArticulo = :codigo AND tmo.Duracion > 0
),
bonos AS (
    SELECT AVG(CAST(ob.TotalMinutos AS float) / NULLIF(ob.TotalPiezas, 0)) AS TiempoMin
    FROM dbo.Ordenes_Bonos_Salidas obs
    INNER JOIN dbo.Ordenes_Articulos oa ON oa.IdArticulo = obs.IdArticulo AND oa.IdOrden = obs.IdOrden
    INNER JOIN dbo.Fases f ON f.IdFase = oa.IdFase AND f.Activa <> 0
    INNER JOIN dbo.Ordenes_Bonos ob ON ob.IdBono = obs.IdBono AND ob.IdOrden = obs.IdOrden
    WHERE ob.TotalPiezas > 0 AND obs.IdArticulo = :codigo
)
-- las dos son agregados sin GROUP BY, asi que cada una devuelve exactamente una
-- fila (con NULL si no hay datos) y el CROSS JOIN da una sola fila
SELECT COALESCE(mo.TiempoMin, b.TiempoMin) AS TiempoMin,
       CASE WHEN mo.TiempoMin IS NULL AND b.TiempoMin IS NOT NULL
            THEN 1 ELSE 0 END AS EsMedio
FROM mano_obra mo CROSS JOIN bonos b
""")


def tiempo_operacion(codigo: str):
    """Tiempo de operacion (min/pieza) del articulo raiz y si es una estimacion.

    Devuelve (minutos, es_medio): minutos es None si no hay ningun dato, y
    es_medio vale 1 cuando el tiempo sale de la media de bonos porque el
    articulo no tiene mano de obra imputada en el ERP."""
    codigo = re.sub(r"[^A-Za-z0-9]", "", str(codigo))
    with get_engine().connect() as cn:
        r = cn.execute(SQL_TIEMPO, {"codigo": codigo}).fetchone()
    if not r or r[0] is None:
        return None, 0
    return float(r[0]), int(r[1] or 0)


SQL_SIN_OPERACION = text("""
SELECT CASE WHEN EXISTS (
    SELECT 1 FROM dbo.Fases_Salidas fs
    INNER JOIN dbo.Fases f ON f.IdFase = fs.IdFase AND f.Activa <> 0
    WHERE fs.IdArticulo = :codigo
      -- mismo criterio que el CTE sin_operacion (ver su comentario): manda
      -- la casilla OrdenTrabajo y nada mas
      AND EXISTS (SELECT 1 FROM dbo.Trabajos_Fases tf
                  WHERE tf.IdFase = fs.IdFase AND tf.OrdenTrabajo = 0)
      AND NOT EXISTS (SELECT 1 FROM dbo.Trabajos_Fases tf
                      WHERE tf.IdFase = fs.IdFase AND tf.OrdenTrabajo = 1)
) THEN 1 ELSE 0 END
""")


def sin_operacion(codigo: str) -> int:
    """1 si la fase activa del articulo no declara operacion de fabricacion
    (todos sus trabajos son 'Sin operacion'). Para el nodo raiz del arbol."""
    codigo = re.sub(r"[^A-Za-z0-9]", "", str(codigo))
    with get_engine().connect() as cn:
        r = cn.execute(SQL_SIN_OPERACION, {"codigo": codigo}).fetchone()
    return int(r[0]) if r else 0


# Precio de compra DEL PROPIO ARTICULO buscado cuando ademas tiene escandallo:
# es el trabajo que se le hace encima (lacado, zincado, remachado...) y se suma
# al coste de sus componentes. Mismo criterio que la rama (c) de EsValorable,
# incluido el salvavidas del CTE 'duplicado'.
#
# El CTE de arriba solo emite una fila para el propio articulo si NO tiene
# escandallo (ver "NIVEL 0 alternativo"). Con escandallo, la raiz nunca aparece
# como fila, asi que la excepcion de tipo 5 de EsValorable no llega a aplicarse
# y el coste del servicio se perdia: buscar 18401007 daba 3,36 EUR (solo su
# chapa) en vez de 3,36 + 4,59 de lacado. Como COMPONENTE de un padre si se
# contaba bien, de modo que el mismo articulo valia distinto segun como se
# mirase. Esto lo recupera para la raiz.
#
# Misma prioridad de fuentes que el CTE 'precio': ultimo ALBARAN, y TARIFA solo
# si no hay ninguna compra (excluyendo la lista del "Proveedor 0").
# OJO con el orden: el freno se aplica AL PRECIO YA ELEGIDO, no a cada candidato.
# Filtrando candidato a candidato, al descartar el precio bueno de 15101013
# (91,00, que coincide con su componente) el TOP 1 se quedaba con un albaran
# viejo de 72,35 que no coincidia, y lo sumaba igual: 163,80 EUR en vez de 91,45.
SQL_SERVICIO_EXTERNO = text("""
WITH candidatos AS (
    SELECT pl.Precio_EURO - pl.Precio_EURO * COALESCE(pl.Descuento, 0) / 100.0 AS Precio,
           1 AS Prioridad, pl.FechaAlbaran AS Fecha, pl.IdPedido AS Desempate
    FROM dbo.Pedidos_Prov_Lineas pl
    WHERE pl.IdArticulo = :codigo
      AND pl.FechaAlbaran IS NOT NULL AND pl.Precio_EURO > 0
      AND COALESCE(pl.Descuento, 0) < 100   -- regalo/unidad gratis, no es precio real (ver CTE 'precio')

    UNION ALL

    SELECT lp.Precio, 2, lp.FechaInsertUpdate, lp.IdLista
    FROM dbo.Listas_Precios_Prov_Art lp
    INNER JOIN dbo.Listas_Precios_Prov lc ON lc.IdLista = lp.IdLista
                                         AND lc.IdProveedor <> '0'
    WHERE lp.IdArticulo = :codigo AND lp.Precio > 0
      AND NOT EXISTS (SELECT 1 FROM dbo.Pedidos_Prov_Lineas pl2
                      WHERE pl2.IdArticulo = :codigo
                        AND pl2.FechaAlbaran IS NOT NULL AND pl2.Precio_EURO > 0
                        AND COALESCE(pl2.Descuento, 0) < 100)
),
elegido AS (
    SELECT TOP 1 Precio FROM candidatos
    ORDER BY Prioridad, Fecha DESC, Desempate DESC
)
SELECT e.Precio FROM elegido e
-- solo si TIENE escandallo (si no, el CTE ya lo valora por la rama "NIVEL 0
-- alternativo" y sumarlo aqui lo contaria dos veces)
-- (la fase tiene que traer COMPONENTES: si esta vacia, el CTE ya valora el
--  articulo por su cuenta y esto sobraria)
WHERE EXISTS (SELECT 1 FROM dbo.Fases_Salidas fs
              INNER JOIN dbo.Fases f ON f.IdFase = fs.IdFase AND f.Activa <> 0
              INNER JOIN dbo.Fases_Entradas fe ON fe.IdFase = f.IdFase
              WHERE fs.IdArticulo = :codigo)
  -- freno: no sumar si ese precio YA esta dentro del escandallo
  AND NOT EXISTS (
        SELECT 1 FROM (
            SELECT fe.IdArticulo AS Hijo
            FROM dbo.Fases f2
            INNER JOIN dbo.Fases_Salidas fs2 ON fs2.IdFase = f2.IdFase AND f2.Activa <> 0
            INNER JOIN dbo.Fases_Entradas fe ON fe.IdFase = f2.IdFase
            WHERE fs2.IdArticulo = :codigo
            UNION ALL
            SELECT ac.IdArticulo FROM dbo.Articulos_Conjuntos ac
            WHERE ac.IdArticuloPadre = :codigo
        ) h
        CROSS APPLY (
            SELECT TOP 1 pl.Precio_EURO AS P
            FROM dbo.Pedidos_Prov_Lineas pl
            WHERE pl.IdArticulo = h.Hijo
              AND pl.FechaAlbaran IS NOT NULL AND pl.Precio_EURO > 0
            ORDER BY pl.FechaAlbaran DESC, pl.IdPedido DESC
        ) x
        WHERE ABS(x.P - e.Precio) <= 0.02 * ABS(e.Precio))
  -- y los revisados a mano (mismo criterio que el CTE 'duplicado')
  AND :codigo NOT IN (/*NO_SUMA_LISTA*/)
""".replace("/*NO_SUMA_LISTA*/", _NO_SUMA_LISTA))


def coste_propio(codigo: str):
    """Precio de compra del propio articulo buscado cuando tiene escandallo: el
    trabajo que se le hace encima. None si no aplica. Se suma a sus materiales."""
    codigo = re.sub(r"[^A-Za-z0-9]", "", str(codigo))
    with get_engine().connect() as cn:
        r = cn.execute(SQL_SERVICIO_EXTERNO, {"codigo": codigo}).fetchone()
    return float(r[0]) if r and r[0] is not None else None


SQL_ESCANDALLO = text("""
SELECT TRY_CONVERT(float, fe.Cantidad) AS Cantidad,
       fe.IdArticulo,
       fe.Descrip AS Descripcion,
       fe.IdTipoUnidad AS Unidad
FROM dbo.Fases f
INNER JOIN dbo.Fases_Salidas  fs ON fs.IdFase = f.IdFase AND f.Activa <> 0
INNER JOIN dbo.Fases_Entradas fe ON fe.IdFase = f.IdFase
WHERE fs.IdArticulo = :codigo
ORDER BY fe.Descrip
""")


def escandallo_directo(codigo: str) -> pd.DataFrame:
    """Componentes DIRECTOS (nivel 0) del articulo, de su fase activa."""
    codigo = re.sub(r"[^A-Za-z0-9]", "", str(codigo))
    with get_engine().connect() as cn:
        return pd.read_sql(SQL_ESCANDALLO, cn, params={"codigo": codigo})


def desglose(codigo: str) -> pd.DataFrame:
    """Devuelve el despiece multinivel del articulo como DataFrame."""
    codigo = re.sub(r"[^A-Za-z0-9]", "", str(codigo))  # sanea el input
    with get_engine().connect() as cn:
        return pd.read_sql(SQL, cn, params={"codigo": codigo})


def articulos_comprados(df: pd.DataFrame) -> pd.DataFrame:
    """Lista plana SOLO de los articulos comprados del escandallo, agregados por
    articulo (suma la cantidad y el coste de todas las ramas donde aparecen)."""
    comprables = df[(df["PrecioCompra"].notna()) | (df["SinPrecio"] == 1)]
    if comprables.empty:
        return pd.DataFrame(columns=["CODIGO", "Descripcion", "Unidad", "Cantidad",
                                     "PrecioCompra", "Descuento", "Coste", "SinPrecio"])
    agg = (comprables.groupby("IdArticulo", as_index=False)
           .agg(Descripcion=("Componente", "first"),
                Tipo=("TipoCompra", "first"),
                Unidad=("Unidad", "first"),
                Cantidad=("Cant", "sum"),
                PrecioCompra=("PrecioCompra", "first"),
                Descuento=("Descuento", "first"),
                Coste=("Coste", "sum"),
                SinPrecio=("SinPrecio", "max"))
           .rename(columns={"IdArticulo": "CODIGO"}))
    return agg.sort_values("Coste", ascending=False, na_position="last")


RATE_OP = 17.0 / 60.0   # 17 EUR/h operario -> EUR/min


def _num(v):
    return None if v is None or (isinstance(v, float) and pd.isna(v)) else float(v)


def _s(v):
    """Campos de texto: NaN -> None. (NaN es un float TRUTHY, asi que colarlo en
    un `x or defecto` cortocircuita el defecto y deja la celda vacia.)"""
    return None if v is None or (isinstance(v, float) and pd.isna(v)) else v


def construir_arbol(df, codigo, nombre, tiempo_raiz=None, sin_op_raiz=0,
                    servicio_raiz=None, medio_raiz=0):
    """Arbol del escandallo con coste de MATERIAL (hojas) y de OPERACION (ramas
    fabricadas = tiempo x 17 EUR/h) y el rollup material+operacion por nodo.

    servicio_raiz: coste del trabajo externo del propio articulo buscado, que
    se suma a sus materiales (ver SQL_SERVICIO_EXTERNO).
    medio_raiz: 1 si el tiempo de la raiz sale de la media de bonos en vez de
    la mano de obra imputada en el ERP."""
    root = {"id": codigo, "nombre": nombre, "cant": 1.0, "unidad": None, "tipo": None,
            "precio": None, "fuente": None, "de_conjunto": 0, "sin_escandallo": 0,
            "sin_operacion": int(sin_op_raiz or 0), "servicio": servicio_raiz,
            "tiempo": tiempo_raiz, "tiempo_medio": int(medio_raiz or 0),
            "coste": servicio_raiz or 0.0, "hijos": []}
    nodos = {f"|{codigo}|": root}

    for r in df.sort_values("Ruta").to_dict(orient="records"):
        ruta = r["Ruta"]
        segs = [s for s in ruta.split("|") if s]
        if segs == [codigo]:
            root["coste"] = _num(r["Coste"]) or 0.0
            root["cant"] = _num(r["Cant"]) or 1.0; root["unidad"] = _s(r["Unidad"])
            root["tipo"] = _s(r["TipoCompra"]); root["precio"] = _num(r["PrecioCompra"])
            root["fuente"] = _s(r.get("PrecioFuente"))
            continue
        nodo = {"id": r["IdArticulo"], "nombre": r["Componente"],
                "cant": _num(r["Cant"]), "unidad": _s(r["Unidad"]),
                "tipo": _s(r["TipoCompra"]), "precio": _num(r["PrecioCompra"]),
                "fuente": _s(r.get("PrecioFuente")),
                "de_conjunto": int(r.get("DeConjunto", 0) or 0),
                "sin_escandallo": int(r.get("SinEscandallo", 0) or 0),
                "sin_operacion": int(r.get("SinOperacion", 0) or 0),
                "tiempo": _num(r.get("TiempoOp")),
                "tiempo_medio": int(r.get("TiempoMedio", 0) or 0),
                "coste": _num(r["Coste"]) or 0.0, "hijos": []}
        nodos[ruta] = nodo
        padre_key = ("|" + "|".join(segs[:-1]) + "|") if len(segs) > 1 else f"|{codigo}|"
        nodos.get(padre_key, root)["hijos"].append(nodo)

    def rollup(n):
        n["es_hoja"] = not n["hijos"]
        # "tiempo" es min/PIEZA; "tiempo_op" son los minutos que consume esta
        # linea (min/pieza x cantidad usada), que es lo que se cobra a 17 €/h.
        if n["es_hoja"]:
            n["coste_op"] = 0.0; n["tiempo_op"] = 0.0
            n["sin_tiempo"] = 0; n["tiempo_efectivo"] = None
        elif n.get("tiempo") is not None and not n.get("tiempo_medio"):
            # mano de obra IMPUTADA en el ERP: declaracion explicita de que la
            # pieza lleva trabajo, manda incluso sobre la casilla de operacion.
            n["tiempo_op"] = round(n["tiempo"] * (n["cant"] or 0), 4)
            n["coste_op"] = round(n["tiempo"] * (n["cant"] or 0) * RATE_OP, 4)
            n["sin_tiempo"] = 0; n["tiempo_efectivo"] = n["tiempo"]
        elif n.get("sin_operacion"):
            # su fase no declara operacion: 0 EUR de mano de obra es CORRECTO,
            # no es un dato que falte -> no se avisa como "sin tiempo".
            # MANDA SOBRE LA MEDIA DE PARTES: ver el comentario largo en app.py
            # (10902103 ANGULO DER. salia a 0,27 min y su espejo a "sin operacion").
            n["coste_op"] = 0.0; n["tiempo_op"] = 0.0
            n["sin_tiempo"] = 0; n["tiempo_efectivo"] = None
        elif n.get("tiempo") is not None:
            # media de los bonos de una pieza que si declara operacion
            n["tiempo_op"] = round(n["tiempo"] * (n["cant"] or 0), 4)
            n["coste_op"] = round(n["tiempo"] * (n["cant"] or 0) * RATE_OP, 4)
            n["sin_tiempo"] = 0; n["tiempo_efectivo"] = n["tiempo"]
        elif n.get("servicio") is not None or n.get("precio") is not None:
            # TRABAJO EXTERNO (lacado, zincado, remachado...): tiene escandallo Y
            # ADEMAS se compra. Lo hace un tercero fuera de GYC, asi que 0 min de
            # mano de obra propia es CORRECTO, no un dato que falte (ver mismo
            # cambio y comentario largo en app.py).
            n["coste_op"] = 0.0; n["tiempo_op"] = 0.0
            n["sin_tiempo"] = 0; n["tiempo_efectivo"] = None
        else:
            n["coste_op"] = 0.0; n["tiempo_op"] = 0.0
            n["sin_tiempo"] = 1; n["tiempo_efectivo"] = None
        mat = n["coste"] or 0.0
        op = n["coste_op"]
        tmin = n["tiempo_op"]
        for h in n["hijos"]:
            rollup(h)
            mat += h["coste_mat"]; op += h["coste_op_total"]; tmin += h["tiempo_op_total"]
        # 6 decimales: con 4, un coste de 8,5e-06 EUR se redondeaba a 0,0 exacto
        # y desaparecia del arbol (ademas de dispararse como "sin coste").
        n["coste_mat"] = round(mat, 6)
        n["coste_op_total"] = round(op, 6)
        # minutos acumulados del nodo: los suyos mas los de todo lo que cuelga
        n["tiempo_op_total"] = round(tmin, 6)
        n["coste_total"] = round(mat + op, 6)

    rollup(root)
    return root


def aplanar_arbol(arbol):
    """Aplana el arbol en orden pre-order: lista de (profundidad, nodo)."""
    filas = []

    def walk(n, depth):
        filas.append((depth, n))
        for h in n["hijos"]:
            walk(h, depth + 1)

    walk(arbol, 0)
    return filas


def _categoria(n):
    """Categoria de una fila para colorear (misma que en la web)."""
    if n["hijos"]:
        return "noesc" if n.get("sin_tiempo") else "rama"
    if n.get("de_conjunto"): return "conjunto"
    if n.get("sin_escandallo"): return "noesc"
    # hueco = no hay precio de albaran, no que falte el tipo ni que el importe
    # sea minimo (el tipo ya no se exige para valorar)
    return "mp" if n.get("precio") is not None else "sinprecio"


_FILLS = {"mp": "FEF3C7", "sinprecio": "FDE0E0", "conjunto": "EDE9FE",
          "noesc": "DBEAFE", "total": "E2E8F0"}


# Los importes se presentan con 4 DECIMALES (precios, costes de material, de
# operacion y totales). El calculo interno sigue con mas precision a proposito:
# redondear el rollup a 4 hacia desaparecer del arbol las lineas de fraccion de
# centimo (ver el comentario de coste_mat en construir_arbol).
_COLS_EUROS = ("Precio €", "Coste MP €", "Coste operación €", "Coste total €",
               "PrecioCompra", "Coste")


def _cuatro_decimales(ws, columnas):
    """Formatea con 4 decimales las columnas de importe de una hoja."""
    for i, nombre in enumerate(columnas, start=1):
        if nombre in _COLS_EUROS:
            for fila in range(2, ws.max_row + 1):
                ws.cell(row=fila, column=i).number_format = "#,##0.0000"


def _resaltar_sin_precio(ws, flags, ncols):
    """Rellena en rojo las filas cuyo flag SinPrecio es 1 (cabecera en fila 1)."""
    rojo = PatternFill("solid", fgColor="FFC7CE")
    for pos, sinp in enumerate(flags):
        if sinp:
            for col in range(1, ncols + 1):
                ws.cell(row=pos + 2, column=col).fill = rojo


def exportar_excel(df: pd.DataFrame, codigo: str, nombre: str, destino) -> None:
    """Escribe el Excel (tres hojas): Desglose (arbol indentado con material y
    operacion, coloreado por categoria), Comprados y Resumen (material+operacion+total)."""
    tiempo_raiz, medio_raiz = tiempo_operacion(codigo)
    arbol = construir_arbol(df, codigo, nombre, tiempo_raiz, sin_operacion(codigo),
                            coste_propio(codigo), medio_raiz)
    filas = aplanar_arbol(arbol)

    # ---- Hoja Desglose: el arbol indentado, con material y operacion ----
    reg, cats = [], []
    for depth, n in filas:
        hoja = not n["hijos"]
        # el articulo raiz puede llevar ademas un trabajo externo propio, que se
        # suma a sus materiales (ver SQL_SERVICIO_EXTERNO)
        servicio = n.get("servicio")
        tipo_txt = ("de conjunto" if n.get("de_conjunto")
                    else "sin escandallo" if n.get("sin_escandallo")
                    else "trabajo sobre la pieza" if servicio is not None
                    else (n.get("tipo") or
                          ("" if hoja
                           else "fabricado sin operación" if n.get("sin_operacion")
                           else "fabricado")))
        reg.append({
            "Nivel": depth,
            "Componente": ("    " * depth) + (n.get("nombre") or ""),
            "IdArticulo": n["id"],
            "Cant": n.get("cant"),
            "Ud": n.get("unidad"),
            "Tipo": tipo_txt,
            "Precio €": n.get("precio") if hoja else servicio,
            # de donde sale el precio: una compra real o la tarifa del proveedor
            "Origen precio": ("tarifa" if n.get("fuente") == "lista"
                              else "albarán" if n.get("fuente") == "albaran"
                              else None) if hoja else None,
            "Coste MP €": n.get("coste") if (hoja or servicio is not None) else None,
            # el tiempo que de verdad se cobra: None si la casilla del ERP dice
            # "sin operacion", aunque el articulo tenga partes fichados
            "Tiempo min": (None if hoja else n.get("tiempo_efectivo")),
            # De donde sale el tiempo. "medio (bonos)" es la MEDIA REAL de lo que
            # se tardo, no un tiempo teorico: el teorico seria el estandar que fija
            # produccion, y ese es justamente el otro caso ("mano de obra").
            "Origen tiempo": (None if (hoja or n.get("tiempo_efectivo") is None)
                              else "medio (bonos)" if n.get("tiempo_medio")
                              else "mano de obra"),
            # minutos acumulados: los de esta linea mas los de todo lo que cuelga
            "Tiempo total min": n.get("tiempo_op_total") or None,
            "Coste operación €": (None if hoja else n.get("coste_op")),
            "Coste total €": n.get("coste_total"),
        })
        cats.append(_categoria(n))

    coste_mat = arbol["coste_mat"]; coste_op = arbol["coste_op_total"]; coste_tot = arbol["coste_total"]
    tiempo_tot = arbol["tiempo_op_total"]
    reg.append({"Nivel": None, "Componente": "TOTAL", "IdArticulo": "", "Cant": None,
                "Ud": "", "Tipo": "", "Precio €": None, "Origen precio": "",
                "Coste MP €": coste_mat, "Tiempo min": None, "Origen tiempo": "",
                "Tiempo total min": tiempo_tot,
                "Coste operación €": coste_op, "Coste total €": coste_tot})
    cats.append("total")
    desg = pd.DataFrame(reg)

    # ---- Comprados y Resumen ----
    val = df[df["Valorado"] == 1]
    coste_mp = round(float(val.loc[val["TipoCompra"] == "MATERIA PRIMA", "Coste"].sum()), 4)
    coste_merc = round(float(val.loc[val["TipoCompra"] == "MERCADERIA", "Coste"].sum()), 4)
    coste_otros = round(coste_mat - coste_mp - coste_merc, 4)

    comprados = articulos_comprados(df)
    comprados_vis = comprados.drop(columns=["SinPrecio"])

    resumen = pd.DataFrame({
        "Concepto": ["Artículo", "Nombre", "Líneas", "Niveles",
                     "Coste MATERIA PRIMA (€)", "Coste MERCADERÍA (€)", "Coste otras compras (€)",
                     "COSTE MATERIAL (€)", "TIEMPO TOTAL (min)", "COSTE OPERACIÓN (€)",
                     "COSTE TOTAL (€)", "Comprables sin precio"],
        "Valor": [codigo, nombre, len(df), int(df["Nivel"].max()),
                  coste_mp, coste_merc, coste_otros,
                  coste_mat, tiempo_tot, coste_op, coste_tot, int(df["SinPrecio"].sum())],
    })

    with pd.ExcelWriter(destino, engine="openpyxl") as xls:
        desg.to_excel(xls, sheet_name="Desglose", index=False)
        comprados_vis.to_excel(xls, sheet_name="Comprados", index=False)
        resumen.to_excel(xls, sheet_name="Resumen", index=False)

        ws = xls.sheets["Desglose"]
        ncols = len(desg.columns)
        fills = {k: PatternFill("solid", fgColor=v) for k, v in _FILLS.items()}
        for i, cat in enumerate(cats):
            fill = fills.get(cat)
            for c in range(1, ncols + 1):
                cell = ws.cell(row=i + 2, column=c)
                if fill:
                    cell.fill = fill
                if cat == "total":
                    cell.font = Font(bold=True)
        for c in range(1, ncols + 1):     # cabecera en negrita
            ws.cell(row=1, column=c).font = Font(bold=True)
        _cuatro_decimales(ws, desg.columns)

        _resaltar_sin_precio(xls.sheets["Comprados"], comprados["SinPrecio"].tolist(), len(comprados_vis.columns))
        _cuatro_decimales(xls.sheets["Comprados"], comprados_vis.columns)


def main():
    ap = argparse.ArgumentParser(description="Despiece multinivel de un articulo del ERP")
    ap.add_argument("codigo", help="Codigo de articulo (ej: 12101021)")
    ap.add_argument("-o", "--output", help="Ruta del Excel de salida")
    args = ap.parse_args()

    df = desglose(args.codigo)
    if df.empty:
        print(f"El articulo '{args.codigo}' no tiene fases activas / desglose.")
        return

    nombre = nombre_articulo(args.codigo)
    ruta = args.output or os.path.join("salidas", f"desglose_{args.codigo}.xlsx")
    os.makedirs(os.path.dirname(os.path.abspath(ruta)), exist_ok=True)
    exportar_excel(df, args.codigo, nombre, ruta)
    print(f"Articulo {args.codigo} ({nombre}): {len(df)} lineas, {df['Nivel'].max()} niveles.")
    print(f"Coste total materia prima: {df['Coste'].sum():.4f} EUR")
    print(f"Excel generado en: {os.path.abspath(ruta)}")


if __name__ == "__main__":
    main()
